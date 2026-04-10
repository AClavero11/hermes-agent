"""Summit 8140 pricing sheet lookup tool.

Local-only tool that reads the Summit 8140 pricing XLSX, merges per-PN data
with recent Jorge Fernandez Gmail history, and exposes a unified lookup.

Privacy wall: tagged ``privacy=high``. Calling the handler while
``HERMES_TOOL_REMOTE_EXEC=1`` raises :class:`PrivacyWallViolation` so the
sheet data never leaves the Mac Studio.

Sheet layout (``'8140 Pricing'`` sheet, headers on row 3, data rows 4+):

    A  PN
    B  Description
    C  A/C (aircraft)
    D  CN (condition — compound, slash-delimited, e.g. ``AR/OH``)
    E  Qty
    F  Kent Ext Cost
    G  May Price (our buy)
    H  Meimin Date
    I  Our Quote (mkt sell)
    J  # Quotes
    K  ILS Sellers
    L  ILS Qty
    M  AAC on ILS
    N  Summit on ILS
    O  ILS OH Price Range
    P  JF Prices
"""

from __future__ import annotations

import logging
import os
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Constants / configuration
# ---------------------------------------------------------------------------

DEFAULT_SHEET_PATH = "/Users/ac/data/summit-8140-pricing.xlsx"
SHEET_NAME = "8140 Pricing"
HEADER_ROW = 3
FIRST_DATA_ROW = 4

# Column index (1-based) -> dataclass field name
COLUMN_MAP = {
    1: "pn",
    2: "description",
    3: "aircraft",
    4: "_cn_raw",
    5: "quantity",
    6: "kent_ext_cost",
    7: "may_buy_price",
    8: "meimin_date",
    9: "our_quote_sell",
    10: "num_quotes",
    11: "ils_sellers",
    12: "ils_qty",
    13: "aac_on_ils",
    14: "summit_on_ils",
    15: "ils_price_range",
    16: "jf_price",
}

JORGE_GMAIL_TOKEN_PATH = Path.home() / ".hermes" / "google_token.json"
# Restrict sender match to Summit MRO to prevent injection from lookalike
# addresses (e.g. jorge.fernandez@attackerdomain.com).
JORGE_SENDER_QUERY = (
    "from:jorge.fernandez@summitmro.com OR from:jfernandez@summitmro.com"
)
# PN sanitizer: keep alphanumerics and dashes only. Aerospace part numbers
# never need other characters, so anything else is either a typo or an
# injection attempt against the Gmail query.
_PN_SAFE_RE = re.compile(r"[^A-Za-z0-9\-]")

# Regex for Summit cost line (case-insensitive, allows commas and decimals)
_SUMMIT_COST_RE = re.compile(
    r"[Ss]ummit\s+[Cc]ost[^\$]*\$\s*([\d,]+(?:\.\d+)?)"
)
# Regex for "over your cost" guidance, case-insensitive
_OVER_YOUR_COST_RE = re.compile(
    r"over your cost[^\d]*([\d,]+(?:\.\d+)?)", re.IGNORECASE
)


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class PrivacyWallViolation(RuntimeError):
    """Raised when a privacy=high task tries to execute in a remote context."""


def assert_local_only() -> None:
    """Block execution if we detect a remote provider context."""
    if os.environ.get("HERMES_TOOL_REMOTE_EXEC") == "1":
        raise PrivacyWallViolation(
            "summit_sheet_lookup is tagged privacy=high and cannot "
            "execute in a remote provider context."
        )


# ---------------------------------------------------------------------------
# Dataclass
# ---------------------------------------------------------------------------


@dataclass
class SummitSheetHit:
    """One row from the Summit 8140 pricing sheet, enriched with overrides."""

    pn: str
    description: Optional[str]
    aircraft: Optional[str]
    conditions: List[str]
    quantity: Optional[float]
    kent_ext_cost: Optional[float]
    may_buy_price: Optional[float]
    meimin_date: Optional[str]
    our_quote_sell: Optional[float]
    num_quotes: Optional[float]
    ils_sellers: Optional[float]
    ils_qty: Optional[float]
    aac_on_ils: Optional[str]
    summit_on_ils: Optional[str]
    ils_price_range: Optional[str]
    jf_price: Optional[float]
    sheet_row_number: int
    cost_basis: Optional[float] = None
    cost_source: Optional[str] = None
    summit_guidance: Optional[str] = None
    guidance_source: Optional[str] = None
    sheet_stale: bool = False
    provenance: Dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Sheet path + cache
# ---------------------------------------------------------------------------


def _get_sheet_path() -> str:
    return os.environ.get("HERMES_SUMMIT_SHEET_PATH", DEFAULT_SHEET_PATH)


class _SheetCache:
    """In-memory cache of the sheet keyed by PN, invalidated on mtime change."""

    def __init__(self) -> None:
        self._path: Optional[str] = None
        self._mtime: Optional[float] = None
        self._rows: Dict[str, SummitSheetHit] = {}

    def _is_stale(self, path: str) -> bool:
        try:
            mtime = os.path.getmtime(path)
        except OSError:
            return True
        return (
            self._path != path
            or self._mtime is None
            or mtime != self._mtime
            or not self._rows
        )

    def load(self, path: Optional[str] = None) -> None:
        path = path or _get_sheet_path()
        if not self._is_stale(path):
            return
        if not os.path.exists(path):
            self._path = path
            self._mtime = None
            self._rows = {}
            return

        try:
            import openpyxl  # local import so tests can stub
        except ImportError as exc:  # pragma: no cover - openpyxl is a runtime dep
            raise RuntimeError(
                "openpyxl is required for summit_sheet_lookup"
            ) from exc

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except (FileNotFoundError, OSError) as exc:
            # TOCTOU: file was removed between the os.path.exists check and
            # this load. Treat as an empty sheet rather than crashing the
            # lookup path.
            logger.warning("Summit sheet load failed for %s: %s", path, exc)
            self._path = path
            self._mtime = None
            self._rows = {}
            return
        try:
            if SHEET_NAME not in wb.sheetnames:
                logger.warning(
                    "Sheet %r not found in %s; available: %s",
                    SHEET_NAME,
                    path,
                    wb.sheetnames,
                )
                rows: Dict[str, SummitSheetHit] = {}
            else:
                ws = wb[SHEET_NAME]
                rows = {}
                for row_idx in range(FIRST_DATA_ROW, ws.max_row + 1):
                    pn_val = ws.cell(row=row_idx, column=1).value
                    if pn_val is None:
                        continue
                    pn = str(pn_val).strip()
                    if not pn:
                        continue
                    raw = {
                        "pn": pn,
                        "description": _clean_str(
                            ws.cell(row=row_idx, column=2).value
                        ),
                        "aircraft": _clean_str(
                            ws.cell(row=row_idx, column=3).value
                        ),
                        "conditions": _expand_conditions(
                            ws.cell(row=row_idx, column=4).value
                        ),
                        "quantity": _to_float(
                            ws.cell(row=row_idx, column=5).value
                        ),
                        "kent_ext_cost": _to_float(
                            ws.cell(row=row_idx, column=6).value
                        ),
                        "may_buy_price": _to_float(
                            ws.cell(row=row_idx, column=7).value
                        ),
                        "meimin_date": _clean_date(
                            ws.cell(row=row_idx, column=8).value
                        ),
                        "our_quote_sell": _to_float(
                            ws.cell(row=row_idx, column=9).value
                        ),
                        "num_quotes": _to_float(
                            ws.cell(row=row_idx, column=10).value
                        ),
                        "ils_sellers": _to_float(
                            ws.cell(row=row_idx, column=11).value
                        ),
                        "ils_qty": _to_float(
                            ws.cell(row=row_idx, column=12).value
                        ),
                        "aac_on_ils": _clean_str(
                            ws.cell(row=row_idx, column=13).value
                        ),
                        "summit_on_ils": _clean_str(
                            ws.cell(row=row_idx, column=14).value
                        ),
                        "ils_price_range": _clean_str(
                            ws.cell(row=row_idx, column=15).value
                        ),
                        "jf_price": _to_float(
                            ws.cell(row=row_idx, column=16).value
                        ),
                        "sheet_row_number": row_idx,
                    }
                    rows[pn] = SummitSheetHit(**raw)
        finally:
            wb.close()

        self._path = path
        try:
            self._mtime = os.path.getmtime(path)
        except OSError:
            self._mtime = None
        self._rows = rows

    def get(self, pn: str) -> Optional[SummitSheetHit]:
        self.load()
        return self._rows.get(pn)

    def reset(self) -> None:
        self._path = None
        self._mtime = None
        self._rows = {}

    def __len__(self) -> int:
        return len(self._rows)


_cache: Optional[_SheetCache] = None


def _get_cache() -> _SheetCache:
    global _cache
    if _cache is None:
        _cache = _SheetCache()
    return _cache


def _reset_cache() -> None:
    """Drop the cached sheet so the next lookup reloads from disk."""
    global _cache
    _cache = None


# ---------------------------------------------------------------------------
# Cell parsing helpers
# ---------------------------------------------------------------------------


def _clean_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _clean_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    return text or None


def _expand_conditions(value: Any) -> List[str]:
    """Split a compound condition code like ``'AR/OH'`` into ``['AR', 'OH']``."""
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    return [piece.strip() for piece in text.split("/") if piece.strip()]


# ---------------------------------------------------------------------------
# Jorge email enrichment
# ---------------------------------------------------------------------------


def _parse_jorge_cost(
    body: str, pn: str
) -> Optional[Tuple[float, str]]:
    """Parse a Jorge email body for a cost figure and a guidance label.

    Returns ``(cost, guidance)`` or ``None`` when nothing usable is found.

    Guidance is one of:
        - ``"70_30"`` if "70/30" appears near the PN reference
        - ``"over_cost"`` if "over your cost" appears
        - ``"other"`` otherwise when a cost is still extracted
    """
    if not body:
        return None

    # Prefer an explicit "Summit cost $X" statement first
    cost: Optional[float] = None
    match = _SUMMIT_COST_RE.search(body)
    if match:
        cost = _to_float(match.group(1))

    if cost is None:
        over_match = _OVER_YOUR_COST_RE.search(body)
        if over_match:
            cost = _to_float(over_match.group(1))

    if cost is None:
        return None

    guidance = "other"
    # "over your cost" wins over 70/30 because it's a harder instruction
    if re.search(r"over your cost", body, re.IGNORECASE):
        guidance = "over_cost"
    else:
        # Look for a 70/30 hint within a small window of the PN mention
        pn_pos = body.find(pn)
        search_slice = body
        if pn_pos != -1:
            start = max(0, pn_pos - 400)
            end = min(len(body), pn_pos + 400)
            search_slice = body[start:end]
        if "70/30" in search_slice:
            guidance = "70_30"

    return cost, guidance


def _fetch_jorge_emails(pn: str) -> List[Dict[str, Any]]:
    """Return Gmail messages from Jorge Fernandez that mention ``pn``.

    Each entry is ``{"message_id": str, "date": "YYYY-MM-DD", "body": str}``.
    Silently returns ``[]`` when Google creds or libraries are unavailable
    — the tool is still useful without enrichment.
    """
    if not JORGE_GMAIL_TOKEN_PATH.exists():
        return []
    try:
        from google.oauth2.credentials import Credentials  # type: ignore
        from googleapiclient.discovery import build  # type: ignore
    except ImportError:
        return []

    try:
        creds = Credentials.from_authorized_user_file(
            str(JORGE_GMAIL_TOKEN_PATH)
        )
        service = build("gmail", "v1", credentials=creds, cache_discovery=False)
        safe_pn = _PN_SAFE_RE.sub("", pn)
        if not safe_pn:
            return []
        query = f'({JORGE_SENDER_QUERY}) "{safe_pn}"'
        resp = (
            service.users()
            .messages()
            .list(userId="me", q=query, maxResults=10)
            .execute()
        )
        results: List[Dict[str, Any]] = []
        for msg_meta in resp.get("messages", []) or []:
            msg_id = msg_meta.get("id")
            if not msg_id:
                continue
            full = (
                service.users()
                .messages()
                .get(userId="me", id=msg_id, format="full")
                .execute()
            )
            body = _extract_email_body(full)
            msg_date = _extract_email_date(full)
            results.append(
                {"message_id": msg_id, "date": msg_date, "body": body}
            )
        return results
    except Exception as exc:  # noqa: BLE001 - network / auth issues shouldn't crash lookup
        logger.debug("Jorge email fetch failed for %s: %s", pn, exc)
        return []


def _extract_email_body(message: Dict[str, Any]) -> str:
    """Best-effort extraction of plain text from a Gmail API message."""
    import base64

    payload = message.get("payload", {}) or {}

    def _walk(part: Dict[str, Any]) -> str:
        mime = part.get("mimeType", "")
        body = part.get("body", {}) or {}
        data = body.get("data")
        if data and mime.startswith("text/"):
            try:
                return base64.urlsafe_b64decode(data.encode("ascii")).decode(
                    "utf-8", errors="replace"
                )
            except Exception:  # noqa: BLE001
                return ""
        collected = ""
        for sub in part.get("parts", []) or []:
            collected += _walk(sub)
        return collected

    return _walk(payload)


def _extract_email_date(message: Dict[str, Any]) -> Optional[str]:
    headers = (message.get("payload", {}) or {}).get("headers", []) or []
    for header in headers:
        if header.get("name", "").lower() == "date":
            raw = header.get("value", "")
            try:
                from email.utils import parsedate_to_datetime

                dt = parsedate_to_datetime(raw)
                if dt is not None:
                    return dt.date().isoformat()
            except Exception:  # noqa: BLE001
                return None
    internal_date = message.get("internalDate")
    if internal_date:
        try:
            ts = int(internal_date) / 1000.0
            return datetime.utcfromtimestamp(ts).date().isoformat()
        except Exception:  # noqa: BLE001
            return None
    return None


# ---------------------------------------------------------------------------
# Merge logic
# ---------------------------------------------------------------------------


def _apply_default_cost_basis(hit: SummitSheetHit) -> None:
    """Set ``cost_basis`` from Kent Ext Cost when no enrichment overrides.

    Also clears every merge-computed field on the cached hit so state
    from a prior lookup does not leak into the current call.
    """
    hit.cost_basis = hit.kent_ext_cost
    hit.cost_source = "kent_ext_cost" if hit.kent_ext_cost is not None else None
    hit.sheet_stale = False
    hit.summit_guidance = None
    hit.guidance_source = None
    hit.provenance = {}


def _merge_jorge_override(
    hit: SummitSheetHit, emails: List[Dict[str, Any]]
) -> None:
    """Override sheet cost with the newest Jorge email newer than Meimin Date."""
    if not emails:
        return

    meimin = hit.meimin_date or ""

    # Sort newest first
    ranked = sorted(
        (
            e
            for e in emails
            if isinstance(e.get("date"), str) and e.get("body")
        ),
        key=lambda e: e["date"],
        reverse=True,
    )

    for email in ranked:
        email_date = email["date"]
        if meimin and email_date <= meimin:
            continue
        parsed = _parse_jorge_cost(email["body"], hit.pn)
        if not parsed:
            continue
        cost, guidance = parsed
        hit.cost_basis = cost
        hit.cost_source = "jorge_email"
        hit.summit_guidance = guidance
        hit.guidance_source = email["message_id"]
        hit.sheet_stale = True
        hit.provenance["jorge_email_id"] = email["message_id"]
        hit.provenance["jorge_email_date"] = email_date
        return


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def check_summit_requirements() -> bool:
    """Return True iff the configured sheet path exists on disk."""
    return os.path.exists(_get_sheet_path())


def summit_sheet_lookup(pn: str) -> Optional[dict]:
    """Look up ``pn`` in the Summit 8140 sheet.

    Returns a dict (``dataclasses.asdict`` form of :class:`SummitSheetHit`)
    or ``None`` if the PN is not present. Jorge email history is merged
    in to override cost_basis when a newer email contradicts the sheet.
    """
    if not pn:
        return None
    hit = _get_cache().get(pn)
    if hit is None:
        return None

    _apply_default_cost_basis(hit)
    try:
        emails = _fetch_jorge_emails(pn)
    except Exception as exc:  # noqa: BLE001
        logger.debug("Jorge enrichment failed for %s: %s", pn, exc)
        emails = []
    _merge_jorge_override(hit, emails)

    return asdict(hit)


def _handle_summit_lookup(args: Dict[str, Any], **_: Any) -> Optional[dict]:
    """Registry handler — enforces privacy wall before any data read."""
    assert_local_only()
    pn = (args or {}).get("pn", "")
    if not isinstance(pn, str):
        pn = str(pn)
    return summit_sheet_lookup(pn.strip())


# ---------------------------------------------------------------------------
# Tool schema + registration
# ---------------------------------------------------------------------------


SUMMIT_SHEET_SCHEMA = {
    "name": "summit_sheet_lookup",
    "description": (
        "Look up a part number in the Summit 8140 pricing sheet with Jorge "
        "Fernandez email enrichment. LOCAL ONLY — data never leaves this "
        "machine."
    ),
    "parameters": {
        "type": "object",
        "properties": {
            "pn": {
                "type": "string",
                "description": "Part number (e.g. '273T1102-8')",
            }
        },
        "required": ["pn"],
    },
}


try:
    from tools.registry import registry

    registry.register(
        name="summit_sheet_lookup",
        toolset="summit_sheet",
        schema=SUMMIT_SHEET_SCHEMA,
        handler=lambda args, **kw: _handle_summit_lookup(args, **kw),
        check_fn=check_summit_requirements,
        description=(
            "Local-only Summit 8140 pricing lookup with Jorge email "
            "enrichment"
        ),
    )
except ImportError:  # pragma: no cover - registry optional for tests
    pass
